#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""자료보유현황.xlsx 생성기.

전국 대학(현재 마스터 351개) 기준으로 "무슨 자료를 갖고 있고 무엇이 비었는지"를
한 장에 모은다. 없는 것은 빈칸으로 두지 않고 반드시 '없음'/'미수집'/'미확인'으로 적는다.

읽기만 한다. 앱 코드·파이프라인 스크립트·DB·PDF 원본을 절대 고치지 않는다.
결과는 이 폴더(inventory/) 안에만 쓴다.

  실행:  python3 build_inventory.py
  캐시 다시 만들기(PDF 재추출):  python3 build_inventory.py --refresh

자료가 늘면 원본 폴더에 파일을 넣고 --refresh 로 다시 돌리면 된다.
"""
import argparse
import json
import os
import re
import sqlite3
import subprocess
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 경로
# ---------------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(HERE, "cache")
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))          # ReBridge_AI공모전용
APP_DATA = os.path.join(ROOT, "Application_main_codes", "src", "data")
PDF_SRC = os.path.join(APP_DATA, "pdf_sources")
V2OUT = os.path.join(ROOT, "data-pipeline", "v2", "out")
DB_PATH = os.path.join(V2OUT, "rebridge.db")
TEXT2028 = os.path.join(V2OUT, "text", "2028")
PLANS = os.path.join(V2OUT, "plans_2028")
OUT_XLSX = os.path.join(HERE, "자료보유현황.xlsx")

NONE = "없음"
FONT = "Arial"          # 한글은 시스템 대체 글꼴로 렌더된다
warnings = []           # 사람이 봐야 할 불확실·실패 항목


def N(s):
    """macOS 파일명은 자모 분리(NFD)로 들어온다. 항상 NFC 로 맞춘다."""
    return unicodedata.normalize("NFC", s or "")


def log(*a):
    print(*a, file=sys.stderr, flush=True)


def warn(msg):
    warnings.append(msg)
    log("  ! " + msg)


# ---------------------------------------------------------------------------
# 0. 캐시 (무거운 PDF 추출)
# ---------------------------------------------------------------------------
def load_cache(refresh=False):
    names = ("ged2027.json", "bigtext.json", "deadline2027.json", "text2025.json")
    need = [f for f in names if not os.path.exists(os.path.join(CACHE, f))]
    if refresh or need:
        log(f"캐시 생성 중 (없는 파일: {need or '전체 갱신'}) — 몇 분 걸린다")
        subprocess.run([sys.executable, os.path.join(HERE, "extract_cache.py")], check=True)
    return tuple(json.load(open(os.path.join(CACHE, f))) for f in names)


# ---------------------------------------------------------------------------
# 1. 대학 이름 매칭
# ---------------------------------------------------------------------------
ABBREV_RULES = [("여자대학교", "여대"), ("교육대학교", "교대"), ("체육대학교", "체대")]


def abbrev_variants(name):
    """자료집마다 약칭 표기가 다르다. 가능한 형태를 모두 만들어 둔다.
    '고려대학교'->'고려대', '서울교육대학교'->'서울교대', '덕성여자대학교'->'덕성여대',
    '한국체육대학교'->'한국체대'/'한국체육대',
    '한국기술교육대학교'->'한국기술교육대'/'한국기술교대'(둘 다 쓰인다)."""
    out = {name}
    for a, b in ABBREV_RULES:
        if a in name:
            out.add(name.replace(a, b))
    out.add(name.replace("대학교", "대"))
    return {s for s in out if s}


def abbrev(name):
    return name.replace("대학교", "대")


def strip_paren(name):
    return re.sub(r"\([^)]*\)", "", name).strip()


def nospace(s):
    return re.sub(r"\s+", "", s)


class Matcher:
    """대학명 -> univId. 자료마다 표기가 달라(전체명/약칭/캠퍼스 괄호) 여러 형태를 받는다."""

    def __init__(self, universities):
        self.univs = universities
        self.by_id = {u["univId"]: u for u in universities}
        self.idx = defaultdict(set)          # 넓은 키(약칭·괄호 제거 포함) -> univId 집합
        self.exact = defaultdict(set)        # 대학명 그대로(괄호 포함) -> univId 집합
        base_groups = defaultdict(list)
        for u in universities:
            nm = u["name"]
            base_groups[strip_paren(nm)].append(u["univId"])
            for k in abbrev_variants(nm):
                self.exact[nospace(k)].add(u["univId"])
            for k in abbrev_variants(nm) | abbrev_variants(strip_paren(nm)):
                self.idx[nospace(k)].add(u["univId"])
        self.base_groups = base_groups

    def lookup(self, raw, region=None):
        """반환: univId 리스트 (0개면 매칭 실패, 2개 이상이면 캠퍼스 분리 대학).
        '동국대학교' 처럼 본교와 캠퍼스 변형이 같은 기본명을 쓰는 경우가 있어
        괄호까지 똑같은 이름(exact)을 먼저 본다."""
        q = nospace(N(raw))
        for table in (self.exact, self.idx):
            for key in (q, nospace(strip_paren(q))):
                if key in table:
                    ids = sorted(table[key])
                    if len(ids) > 1 and region:
                        narrowed = [i for i in ids if self.by_id[i].get("region") == region]
                        if narrowed:
                            return narrowed
                    return ids
        return []

    def search_patterns(self, univ_id):
        """대형 자료집 전문에서 이 대학이 언급됐는지 볼 정규식들.
        자료집은 '고려대', '한양대(ERICA)' 같은 약칭을 쓴다. 캠퍼스 형제가 있으면
        형제의 괄호 표기를 부정탐색으로 빼서 본교/분교를 갈라 본다."""
        u = self.by_id[univ_id]
        nm = u["name"]
        base = strip_paren(nm)
        label = re.search(r"\(([^)]*)\)", nm)
        siblings = [self.by_id[i]["name"] for i in self.base_groups[base] if i != univ_id]
        sib_labels = [m.group(1) for m in
                      (re.search(r"\(([^)]*)\)", s) for s in siblings) if m]
        pats = []
        neg = ""
        if not label and sib_labels:
            # 본교 항목: 형제 캠퍼스 표기가 뒤따르면 그건 다른 학교다
            neg = r"(?!\(" + "|".join(re.escape(l) for l in sib_labels) + r"\))"
        for form in sorted(abbrev_variants(nm)):
            pats.append(re.escape(nospace(form)) + neg)
        return [re.compile(p) for p in pats]


# ---------------------------------------------------------------------------
# 2. 원천별 수집
# ---------------------------------------------------------------------------
def collect():
    d = {}

    # --- 대학 마스터 ------------------------------------------------------
    universities = json.load(open(os.path.join(APP_DATA, "universities.json")))
    d["universities"] = universities
    M = Matcher(universities)
    d["matcher"] = M
    log(f"대학 마스터 {len(universities)}개")

    # --- 앱 JSON ----------------------------------------------------------
    admissions = json.load(open(os.path.join(APP_DATA, "admissions.json")))
    comparative = json.load(open(os.path.join(APP_DATA, "comparative_2028.json")))
    cutlines = json.load(open(os.path.join(APP_DATA, "cutlines_2025.json")))
    d["admissions"], d["comparative"], d["cutlines"] = admissions, comparative, cutlines

    adm_by_univ = defaultdict(list)
    for a in admissions:
        adm_by_univ[a["univId"]].append(a)
    d["adm_by_univ"] = adm_by_univ

    # --- DB ---------------------------------------------------------------
    db = {"ok": False, "std_code": {}, "cutline_rows": {}, "adm_prog": {},
          "tables": []}
    if os.path.exists(DB_PATH):
        try:
            con = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
            cur = con.cursor()
            db["tables"] = [r[0] for r in cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table'")]
            if "university" in db["tables"]:
                for uid, code in cur.execute("SELECT univ_id, std_code FROM university"):
                    db["std_code"][uid] = code
            if "cutline" in db["tables"]:
                for uid, n in cur.execute(
                        "SELECT univ_id, COUNT(*) FROM cutline WHERE year=2025 GROUP BY univ_id"):
                    db["cutline_rows"][uid] = n
            if "admission_program" in db["tables"]:
                for uid, n in cur.execute(
                        "SELECT a.univ_id, COUNT(*) FROM admission_program ap "
                        "JOIN admission a USING(admission_id) GROUP BY a.univ_id"):
                    db["adm_prog"][uid] = n
            db["ok"] = True
            con.close()
        except Exception as e:                                  # noqa: BLE001
            warn(f"DB 읽기 실패({e}) — DB 기반 열은 'DB 없음'으로 채운다")
    else:
        warn("rebridge.db 를 찾지 못했다 — DB 기반 열은 'DB 없음'")
    d["db"] = db

    # --- 2028 시행계획 PDF / hwp -----------------------------------------
    pdf2028 = defaultdict(list)      # univId -> [파일명]
    unmatched_pdf = []
    manifest_path = os.path.join(TEXT2028, "_manifest.json")
    manifest = json.load(open(manifest_path)) if os.path.exists(manifest_path) else {"items": []}
    file2uid = {N(i["file"]): i["univId"] for i in manifest.get("items", [])}
    man_pages = {N(i["file"]): i.get("pages") for i in manifest.get("items", [])}
    dir2028 = os.path.join(PDF_SRC, "2028")
    files2028 = sorted(N(f) for f in os.listdir(dir2028) if f.lower().endswith(".pdf"))
    for f in files2028:
        uid = file2uid.get(f)
        if not uid:                                  # 텍스트 추출이 안 된 파일 -> 이름으로 매칭
            ids = M.lookup(re.split(r"\[", f)[0])
            uid = ids[0] if len(ids) == 1 else (ids[0] if ids else None)
            if not uid:
                unmatched_pdf.append(f)
                continue
        pdf2028[uid].append(f)
    d["pdf2028"], d["files2028"], d["man_pages"] = pdf2028, files2028, man_pages
    d["unmatched_pdf"] = unmatched_pdf
    d["no_text_2028"] = sorted(set(files2028) - set(file2uid))
    if d["no_text_2028"]:
        warn(f"2028 시행계획 PDF 중 텍스트 추출 결과가 없는 파일 {len(d['no_text_2028'])}개: "
             + ", ".join(d["no_text_2028"]))
    if unmatched_pdf:
        warn(f"2028 PDF 대학명 매칭 실패 {len(unmatched_pdf)}건: {unmatched_pdf}")

    hwp2028 = defaultdict(list)
    dirhwp = os.path.join(PDF_SRC, "2028_hwp")
    files_hwp = sorted(N(f) for f in os.listdir(dirhwp)
                       if f.lower().endswith((".hwp", ".hwpx")))
    for f in files_hwp:
        ids = M.lookup(re.split(r"\[", f)[0])
        if ids:
            hwp2028[ids[0]].append(f)
        else:
            warn(f"2028 hwp 대학명 매칭 실패: {f}")
    d["hwp2028"], d["files_hwp"] = hwp2028, files_hwp

    # 텍스트 추출 결과 (univId 별 jsonl)
    text_files = defaultdict(list)
    text_body = {}
    if os.path.isdir(TEXT2028):
        for f in sorted(os.listdir(TEXT2028)):
            if not f.endswith(".jsonl"):
                continue
            uid = N(f)[:-6].split("_")[0]
            text_files[uid].append(N(f))
    d["text_files"] = text_files

    # 2028 시행계획 텍스트에서 '정시 수능 영역별 반영비율' 흔적 찾기
    csat_ratio = {}
    for uid, fl in text_files.items():
        hit = False
        for f in fl:
            try:
                body = "".join(json.loads(l).get("text", "")
                               for l in open(os.path.join(TEXT2028, f)))
            except Exception:                                    # noqa: BLE001
                continue
            t = nospace(body)
            if "영역별반영비율" in t or "수능영역별" in t or "수능반영비율" in t:
                hit = True
                break
        csat_ratio[uid] = hit
    d["csat_ratio"] = csat_ratio

    # --- 2028 판정 산출물 --------------------------------------------------
    def read_jsonl(path):
        rows = []
        if not os.path.exists(path):
            warn(f"{os.path.basename(path)} 없음")
            return rows
        for line in open(path):
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except Exception:                                    # noqa: BLE001
                pass
        return rows

    ged_rows = read_jsonl(os.path.join(PLANS, "ged.jsonl"))
    conv_rows = read_jsonl(os.path.join(PLANS, "conversion.jsonl"))
    d["ged_rows"], d["conv_rows"] = ged_rows, conv_rows

    ged_by_uid = defaultdict(list)
    for r in ged_rows:
        ged_by_uid[r["univId"]].append(r)
    conv_by_uid = defaultdict(list)
    for r in conv_rows:
        conv_by_uid[r["univId"]].append(r)
    d["ged_by_uid"], d["conv_by_uid"] = ged_by_uid, conv_by_uid

    # --- 캐시: 검정고시 지원가능 전형(2027, 5권역) -------------------------
    ged2027, bigtext, deadline_rows, text2025 = load_cache(args.refresh)
    d["ged2027_raw"], d["bigtext"], d["text2025"] = ged2027, bigtext, text2025

    REGION_LABEL = {
        "2027_검정고시_지원가능전형_수도권_0821.pdf": "수도권",
        "2027_검정고시_지원가능전형_중부권_0812.pdf": "중부권",
        "2027_검정고시_지원가능전형_영남권_0812.pdf": "영남권",
        "2027_검정고시_지원가능전형_호남권_0812.pdf": "호남권",
        "2027_검정고시_지원가능전형_강원제주권_0812.pdf": "강원제주권",
    }
    d["REGION_LABEL"] = REGION_LABEL
    ged27_count = Counter()          # univId -> 전형 수
    ged27_zone = {}                  # univId -> 권역
    ged27_unmatched = Counter()
    zone_stats = defaultdict(lambda: {"rows": 0, "univs": set()})
    for r in ged2027["rows"]:
        zone = REGION_LABEL.get(N(r["file"]), N(r["file"]))
        zone_stats[zone]["rows"] += 1
        zone_stats[zone]["univs"].add(r["univ"])
        ids = M.lookup(r["univ"], r.get("region"))
        if not ids:
            ged27_unmatched[r["univ"]] += 1
            continue
        for uid in ids:
            ged27_count[uid] += 1
            ged27_zone[uid] = zone
    d["ged27_count"], d["ged27_zone"] = ged27_count, ged27_zone
    d["zone_stats"] = {k: {"rows": v["rows"], "univs": len(v["univs"])}
                       for k, v in zone_stats.items()}
    d["ged27_names"] = sorted({r["univ"] for r in ged2027["rows"]})
    if ged27_unmatched:
        warn("5권역 PDF 대학명 매칭 실패: " + str(dict(ged27_unmatched)))
    d["ged27_unmatched"] = ged27_unmatched

    # --- 캐시: 2027 수시 접수마감 -----------------------------------------
    d["deadline"], d["deadline_unmatched"] = parse_deadlines(deadline_rows, M)

    # --- 캐시: 대형 자료집 수록 여부 ---------------------------------------
    d["in_analysis"] = presence(bigtext.get("kcue2027_analysis") or "", M, universities)
    d["in_119"] = presence(bigtext.get("kcue2027_119") or "", M, universities)

    # 2028 설명회는 이미지 PDF라 본문 텍스트가 없다. 목차만 읽힌다.
    brief_txt = bigtext.get("kcue2028_briefing") or ""
    brief_names = sorted(set(
        re.findall(r"^\s*\d+\.\s*([가-힣A-Za-z()]+대학교)", brief_txt, re.M)))
    d["brief_names"] = brief_names
    d["in_briefing"] = set()
    for nm in brief_names:
        for uid in M.lookup(nm):
            d["in_briefing"].add(uid)
    if len(brief_txt) < 30000:
        warn(f"2028 대입정보 설명회 PDF는 이미지 기반이라 본문 텍스트가 거의 없다"
             f"(추출 {len(brief_txt)}자/184쪽). 목차의 제3부 대학 {len(brief_names)}개만 판정했다")

    # 2028 반영과목 자료집: 서두에 발표/미지정/제외 대학 명단이 그대로 있다
    d["subjects"] = parse_subject_lists(bigtext.get("kcue2028_subjects") or "", M)

    # --- 2025 대교협 전형결과 지역별 PDF ----------------------------------
    d["in_2025"], d["thin_2025"] = presence_2025(text2025, M, universities)
    if d["thin_2025"]:
        warn(f"2025 전형결과 PDF {len(d['thin_2025'])}개가 스캔 이미지라 글자가 거의 안 뽑힌다"
             f"(OCR 필요). 그 지역 대학은 '2025 대교협 자료집 수록'이 실제보다 낮게 나온다: "
             + ", ".join(d["thin_2025"]))

    return d


def presence(text, matcher, universities):
    """대형 자료집 전문에 대학명이 등장하는지. 대략적인 '수록 여부' 판정."""
    t = nospace(N(text))
    found = set()
    if not t:
        return found
    for u in universities:
        for pat in matcher.search_patterns(u["univId"]):
            if pat.search(t):
                found.add(u["univId"])
                break
    return found


def presence_2025(text2025, matcher, universities):
    """2025 전형결과 PDF(지역별 34개) 전문에 대학명이 등장하는지 + 텍스트가 빈약한 파일."""
    joined = nospace(N("\n".join(text2025.values())))
    found = set()
    for u in universities:
        for pat in matcher.search_patterns(u["univId"]):
            if pat.search(joined):
                found.add(u["univId"])
                break
    thin = sorted(f for f, t in text2025.items() if len(t) < 5000)
    return found, thin


def parse_subject_lists(text, matcher):
    """반영과목 자료집 서두의 세 명단(발표 / 권장과목 미지정 / 학종 미실시 제외)."""
    out = {"발표": set(), "미지정": set(), "제외": set(), "raw": {}}
    lines = [l.strip() for l in N(text).splitlines()]
    bucket = None
    keys = {"모집단위별 반영과목 및 권장과목 발표 대학": "발표",
            "권장 이수과목 미지정 대학": "미지정",
            "학생부종합전형 미실시에 따른 자료집 제외 대학": "제외"}
    names = defaultdict(list)
    for l in lines:
        hit = next((v for k, v in keys.items() if k in l), None)
        if hit:
            bucket = hit
            continue
        if bucket:
            if l.startswith("※") or l.startswith("CONTENTS"):
                bucket = None
                continue
            # -layout 텍스트는 한 줄에 여러 대학이 공백으로 나열된다
            for tok in re.split(r"\s{1,}", l):
                tok = tok.strip()
                if re.fullmatch(r"[가-힣A-Za-z·]+(대학교|대학)(\([^)]+\))?", tok):
                    names[bucket].append(tok)
    for b, ns in names.items():
        out["raw"][b] = ns
        for nm in ns:
            for uid in matcher.lookup(nm):
                out[b].add(uid)
    return out


def parse_deadlines(rows, matcher):
    """캐시된 접수마감 표 -> univId 별 '9.11.(금) 18:00'.
    일자 칸은 같은 지역 안에서 이어지므로 비어 있으면 위 값을 가져온다."""
    out, unmatched = {}, Counter()
    cur_date = ""
    for r in rows:
        date = N(r.get("date") or "").strip() or cur_date
        cur_date = date
        time_ = N(r.get("time") or "").strip()
        names = N(r.get("univs") or "")
        if not names:
            continue
        for nm in names.split(","):
            nm = nm.strip().strip("･").strip()
            if not nm or "대" not in nm:
                continue
            ids = matcher.lookup(nm)
            if not ids:
                unmatched[nm] += 1
                continue
            stamp = f"{date} {time_}" if time_ and time_ != "-" else f"{date} (시각 미표기)"
            for uid in ids:
                out[uid] = stamp
    if unmatched:
        warn("접수마감 PDF 대학명 매칭 실패: " + str(dict(unmatched)))
    return out, unmatched


# ---------------------------------------------------------------------------
# 3. 대학별 행 만들기
# ---------------------------------------------------------------------------
def yn(flag):
    return "있음" if flag else NONE


def build_rows(d):
    M = d["matcher"]
    db = d["db"]
    rows = []
    guides2026 = set()
    gdir = os.path.join(PDF_SRC, "guides_2026")
    if os.path.isdir(gdir):
        for name in sorted(os.listdir(gdir)):
            if name.startswith("."):
                continue
            for uid in M.lookup(N(name)):
                guides2026.add(uid)
    d["guides2026"] = guides2026

    for u in d["universities"]:
        uid = u["univId"]
        r = {}
        r["univId"] = uid
        r["대학명"] = u["name"]
        r["지역"] = u.get("region") or NONE
        r["설립"] = u.get("establishment") or NONE
        r["유형"] = u.get("kind") or NONE
        r["입학처 URL"] = u.get("admissionOfficeUrl") or NONE
        code = db["std_code"].get(uid) if db["ok"] else None
        r["표준코드"] = code if code else ("없음" if db["ok"] else "DB 없음")

        # ---- 2028 시행계획
        pdfs = d["pdf2028"].get(uid, [])
        r["2028 PDF 보유(파일명)"] = " / ".join(pdfs) if pdfs else NONE
        hwps = d["hwp2028"].get(uid, [])
        r["2028 hwp 원본"] = " / ".join(hwps) if hwps else NONE
        tfs = d["text_files"].get(uid, [])
        if tfs:
            pages = sum(d["man_pages"].get(f, 0) or 0 for f in pdfs)
            r["2028 텍스트 추출"] = f"있음({len(tfs)}건, {pages}쪽)" if pages else f"있음({len(tfs)}건)"
        else:
            r["2028 텍스트 추출"] = NONE

        grs = d["ged_by_uid"].get(uid, [])
        if grs:
            vals = {g["eligible"] for g in grs}
            for pref in ("가능", "조건부", "확인필요", "불가", "판정불가"):
                if pref in vals:
                    label = "판정불가(근거 문구 못 찾음)" if pref == "판정불가" else pref
                    r["2028 검정고시 판정"] = label if len(vals) == 1 else f"{label} / 캠퍼스별 상이"
                    break
        elif pdfs:
            r["2028 검정고시 판정"] = "미처리"
        else:
            r["2028 검정고시 판정"] = "판정불가(원문 없음)"

        comp = d["comparative"].get(uid)
        convs = d["conv_by_uid"].get(uid, [])
        if comp and "conversion" in comp:
            ok = any(c.get("monotonic") for c in convs) if convs else True
            r["2028 환산표"] = "구조화·검증통과" if ok else "구조화·검증실패"
        elif convs:
            r["2028 환산표"] = ("구조화·검증실패" if not any(c.get("monotonic") for c in convs)
                             else "구조화·앱 미반영")
        elif comp and (comp.get("comparativeGrade") or "").strip():
            r["2028 환산표"] = "원문만"
        else:
            r["2028 환산표"] = NONE

        has_quote = any((g.get("quote") or "").strip() for g in grs)
        has_raw = bool(comp and (comp.get("comparativeGrade") or "").strip())
        r["2028 원문 게시용 텍스트 준비"] = yn(has_quote or has_raw)

        r["2028 설명회 수록"] = yn(uid in d["in_briefing"])
        subj = d["subjects"]
        if uid in subj["발표"]:
            r["2028 반영과목 자료집 수록"] = "있음"
        elif uid in subj["미지정"]:
            r["2028 반영과목 자료집 수록"] = "권장과목 미지정"
        elif uid in subj["제외"]:
            r["2028 반영과목 자료집 수록"] = "자료집 제외(학종 미실시)"
        else:
            r["2028 반영과목 자료집 수록"] = NONE

        # ---- 2027
        r["2027 시행계획"] = NONE
        r["2027 검정고시 지원가능 전형 수록(권역)"] = d["ged27_zone"].get(uid, NONE)
        r["2027 수록 전형 수"] = d["ged27_count"].get(uid, 0)
        r["2027 수시 모집요강"] = NONE
        r["2027 정시 모집요강"] = NONE
        r["2027 전형분석 자료집 수록"] = yn(uid in d["in_analysis"])
        r["2027 대입정보119 수록"] = yn(uid in d["in_119"])
        r["2027 수시 접수마감 일시"] = d["deadline"].get(uid, NONE)

        # ---- 2026
        r["2026 전형결과"] = NONE
        r["2026 수시 모집요강"] = yn(uid in guides2026)

        # ---- 2025 결과
        r["2025 대교협 자료집 수록"] = yn(uid in d["in_2025"])
        cl = d["cutlines"].get(uid)
        r["2025 합격선 블록 수"] = len(cl) if cl else 0
        r["2025 학과 행 수(DB)"] = (db["cutline_rows"].get(uid, 0) if db["ok"] else "DB 없음")

        # ---- 앱 현재
        adm = d["adm_by_univ"].get(uid, [])
        n_conf = sum(1 for a in adm if str(a.get("status", "")).startswith("confirmed"))
        n_base = sum(1 for a in adm if a.get("status") == "baseline")
        r["앱 전형 행 수"] = len(adm)
        r["앱 확인(confirmed)"] = n_conf
        r["앱 템플릿(baseline)만"] = n_base
        r["앱 합격선 있음"] = yn(bool(cl))
        app_conv = bool(comp and "conversion" in comp)
        r["앱 공식 환산표 반영"] = yn(app_conv)
        r["앱 실제 계산 가능(둘 다)"] = yn(app_conv and bool(cl))

        # ---- 추가 4열
        if d["csat_ratio"].get(uid):
            r["정시 수능 반영비율 보유"] = "원문만"
        elif tfs:
            r["정시 수능 반영비율 보유"] = NONE
        else:
            r["정시 수능 반영비율 보유"] = "없음(원문 텍스트 자체 없음)"

        n_rc = sum(1 for a in adm if a.get("recruitCount") is not None)
        ap = db["adm_prog"].get(uid, 0) if db["ok"] else None
        if not adm:
            r["모집인원 확보"] = NONE
        elif n_rc == 0:
            r["모집인원 확보"] = f"없음 (0/{len(adm)}, DB {('%d행' % ap) if ap is not None else '없음'})"
        else:
            r["모집인원 확보"] = f"{n_rc}/{len(adm)} (DB {('%d행' % ap) if ap is not None else '없음'})"

        r["검정고시 출신 입학생 수"] = "미수집"
        r["대학 자체 입시결과 페이지 URL"] = "미확인"

        # ---- 비고
        notes = []
        if pdfs and not tfs:
            notes.append("2028 PDF는 있으나 텍스트 추출 결과 없음")
        if len(pdfs) > 1:
            notes.append(f"캠퍼스 PDF {len(pdfs)}건을 이 univId로 합침")
        if hwps and not pdfs:
            notes.append("hwp 원본만 있고 PDF 없음")
        in_lists = [k for k in ("발표", "미지정", "제외") if uid in subj[k]]
        if len(in_lists) > 1:
            notes.append("반영과목 자료집 명단에 중복 등재(" + "·".join(in_lists) + ") — 원문이 그렇다")
        r["비고"] = " · ".join(notes) if notes else "-"
        rows.append(r)

    # 마스터에 없는 대학(5권역 PDF·2028 PDF 기준) 추가
    extra = []
    for nm in d["ged27_unmatched"]:
        extra.append((nm, "5권역 PDF"))
    for f in d.get("unmatched_pdf", []):
        extra.append((re.split(r"\[", f)[0], "2028 시행계획 PDF"))
    for nm, src in extra:
        blank = {k: NONE for k in rows[0]}
        blank["univId"] = ""
        blank["대학명"] = nm
        blank["비고"] = f"마스터에 없음 (출처: {src})"
        blank["검정고시 출신 입학생 수"] = "미수집"
        blank["대학 자체 입시결과 페이지 URL"] = "미확인"
        for k in ("2027 수록 전형 수", "2025 합격선 블록 수", "앱 전형 행 수",
                  "앱 확인(confirmed)", "앱 템플릿(baseline)만"):
            blank[k] = 0
        rows.append(blank)
    d["extra_rows"] = extra
    return rows


# ---------------------------------------------------------------------------
# 4. 자료 목록
# ---------------------------------------------------------------------------
def pdf_pages(path):
    try:
        r = subprocess.run(["pdfinfo", path], capture_output=True, text=True, timeout=60)
        m = re.search(r"^Pages:\s+(\d+)", r.stdout, re.M)
        return int(m.group(1)) if m else ""
    except Exception:                                            # noqa: BLE001
        return ""


def mtime(path):
    try:
        return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d")
    except OSError:
        return ""


def build_docs(d, rows):
    M = d["matcher"]
    docs = []
    KCUE = "한국대학교육협의회"

    # 2028 시행계획 PDF 216개 — 한 줄로 묶는다
    dir2028 = os.path.join(PDF_SRC, "2028")
    pages_total = sum(v or 0 for v in d["man_pages"].values())
    n_text = len(d["text_files"])
    docs.append(dict(
        자료명=f"2028학년도 대학입학전형 시행계획(1차수) PDF {len(d['files2028'])}개",
        학년도="2028", 발행처="각 대학(대교협 취합)",
        경로=os.path.relpath(dir2028, ROOT), 쪽수=f"{pages_total}(추출분 합계)", 형식="PDF",
        텍스트=f"가능 {len(d['files2028']) - len(d['no_text_2028'])}개 / 실패 {len(d['no_text_2028'])}개",
        수록대학=len(d["pdf2028"]), 입수일=mtime(dir2028), 출처URL="",
        용도="검정고시 지원자격·환산표·전형방법의 1차 근거",
        비고=("텍스트 추출 실패: " + ", ".join(d["no_text_2028"])) if d["no_text_2028"] else ""))

    # 2028 hwp 원본
    for f in d["files_hwp"]:
        p = os.path.join(PDF_SRC, "2028_hwp", f)
        ids = M.lookup(re.split(r"\[", f)[0])
        docs.append(dict(
            자료명=f, 학년도="2028", 발행처="각 대학",
            경로=os.path.relpath(p, ROOT), 쪽수="-",
            형식="HWPX" if f.lower().endswith(".hwpx") else "HWP",
            텍스트="가능(hwp5txt / hwpx=zip 내 XML)", 수록대학=len(ids) or 0,
            입수일=mtime(p), 출처URL="", 용도="PDF가 없거나 표가 깨진 대학의 원본 대체",
            비고="" if ids else "대학명 매칭 실패"))

    # 2025 대교협 전형결과
    t2025 = d["text2025"]
    for f in sorted(t2025):
        p = os.path.join(PDF_SRC, "2025", f)
        txt = t2025[f]
        flat = nospace(N(txt))
        n_univ = sum(1 for u in d["universities"]
                     if any(pat.search(flat) for pat in M.search_patterns(u["univId"])))
        docs.append(dict(
            자료명=f, 학년도="2025", 발행처=KCUE, 경로=os.path.relpath(p, ROOT),
            쪽수=pdf_pages(p), 형식="PDF",
            텍스트="가능" if len(txt) >= 5000 else f"부분/불가({len(txt)}자·이미지 추정)",
            수록대학=n_univ, 입수일=mtime(p), 출처URL="",
            용도="2025 합격선(70%컷 등) 원천", 비고=""))

    # 검정고시 지원가능 전형 5권역
    for f, st in d["ged2027_raw"]["byRegionFile"].items():
        p = os.path.join(PDF_SRC, "ged_eligible_2027", f)
        zone = d["REGION_LABEL"].get(N(f), N(f))
        n_univ = len({r["univ"] for r in d["ged2027_raw"]["rows"] if N(r["file"]) == N(f)})
        docs.append(dict(
            자료명=N(f), 학년도="2027", 발행처=KCUE, 경로=os.path.relpath(p, ROOT),
            쪽수=st["pages"], 형식="PDF(표)", 텍스트=f"가능(표 추출 {st['pages']}/{st['pages']}쪽 성공)",
            수록대학=n_univ, 입수일=mtime(p), 출처URL="",
            용도=f"{zone} 검정고시 지원가능 전형 목록(전형 {st['rows']}건)", 비고=""))

    # kcue 2027
    kcue2027 = [
        ("2027학년도 수시모집 지역별 전형분석 자료집.pdf", "지역별 전형일정·전형방법·변경사항",
         len(d["in_analysis"])),
        ("2027학년도 대입정보 119 자료집(260112).pdf", "전형 전반 해설", len(d["in_119"])),
        ("2027학년도 대입정보 119 신구대조표(260112).pdf", "전년 대비 변경 대조", "미판정"),
        ("2027학년도 수시모집 전형 일정 (접수 마감).pdf", "대학별 원서접수 마감 일시",
         len(d["deadline"])),
        ("2027학년도_수시모집_주요사항_안내(26.08.04))_수정.hwp", "수시 주요사항", "미판정"),
    ]
    for f, use, n_univ in kcue2027:
        p = os.path.join(PDF_SRC, "kcue_2027", f)
        is_pdf = f.lower().endswith(".pdf")
        docs.append(dict(
            자료명=f, 학년도="2027", 발행처=KCUE, 경로=os.path.relpath(p, ROOT),
            쪽수=pdf_pages(p) if is_pdf else "-", 형식="PDF" if is_pdf else "HWP",
            텍스트="가능" if is_pdf else "가능(hwp5txt)", 수록대학=n_univ,
            입수일=mtime(p), 출처URL="", 용도=use,
            비고="" if n_univ != "미판정" else "대학별 수록 여부는 판정하지 않았다"))

    # kcue 2028
    p = os.path.join(PDF_SRC, "kcue_2028", "2028 대입 정보 설명회_최종_압축.pdf")
    docs.append(dict(
        자료명=os.path.basename(p), 학년도="2028", 발행처=KCUE,
        경로=os.path.relpath(p, ROOT), 쪽수=pdf_pages(p), 형식="PDF(이미지)",
        텍스트=f"부분/불가(전체 {len(d['bigtext'].get('kcue2028_briefing') or '')}자만 추출)",
        수록대학=len(d["in_briefing"]), 입수일=mtime(p), 출처URL="",
        용도="제3부 대학별 2028 전형 운영계획",
        비고="본문이 이미지라 목차의 7개 대학만 판정. 필요하면 OCR 필요"))
    p = os.path.join(PDF_SRC, "kcue_2028", "2028 모집단위별 반영과목 및 대학별 권장과목 자료집.pdf")
    docs.append(dict(
        자료명=os.path.basename(p), 학년도="2028", 발행처=KCUE,
        경로=os.path.relpath(p, ROOT), 쪽수=pdf_pages(p), 형식="PDF",
        텍스트="가능", 수록대학=len(d["subjects"]["발표"]), 입수일=mtime(p), 출처URL="",
        용도="모집단위별 반영과목·대학별 권장과목",
        비고=f"발표 {len(d['subjects']['발표'])} / 권장과목 미지정 {len(d['subjects']['미지정'])} "
            f"/ 학종 미실시 제외 {len(d['subjects']['제외'])}"))

    # guides_2026
    gdir = os.path.join(PDF_SRC, "guides_2026")
    if os.path.isdir(gdir):
        for name in sorted(os.listdir(gdir)):
            sub = os.path.join(gdir, name)
            if not os.path.isdir(sub):
                continue
            for f in sorted(os.listdir(sub)):
                if f.startswith("."):
                    continue
                p = os.path.join(sub, f)
                docs.append(dict(
                    자료명=f"{N(name)} / {N(f)}", 학년도="2026", 발행처=N(name),
                    경로=os.path.relpath(p, ROOT), 쪽수=pdf_pages(p), 형식="PDF",
                    텍스트="가능", 수록대학=1, 입수일=mtime(p), 출처URL="",
                    용도="2026 수시 모집요강(표본 2개교뿐)", 비고=""))

    # 파이프라인 산출물
    for path, name, use, n in [
        (DB_PATH, "rebridge.db (SQLite)", "university/admission/cutline/ged_* 통합 DB",
         len(d["db"]["tables"])),
        (os.path.join(PLANS, "ged.jsonl"), "plans_2028/ged.jsonl",
         "2028 시행계획 PDF 기반 검정고시 지원가능 판정", len(d["ged_by_uid"])),
        (os.path.join(PLANS, "conversion.jsonl"), "plans_2028/conversion.jsonl",
         "2028 시행계획 PDF 기반 검정고시 성적 환산표", len(d["conv_by_uid"])),
        (os.path.join(APP_DATA, "admissions.json"), "src/data/admissions.json",
         "앱이 실제로 쓰는 전형 목록", len(d["adm_by_univ"])),
        (os.path.join(APP_DATA, "comparative_2028.json"), "src/data/comparative_2028.json",
         "앱이 쓰는 환산표/비교내신 원문", len(d["comparative"])),
        (os.path.join(APP_DATA, "cutlines_2025.json"), "src/data/cutlines_2025.json",
         "앱이 쓰는 2025 합격선", len(d["cutlines"])),
        (os.path.join(APP_DATA, "universities.json"), "src/data/universities.json",
         "대학 마스터", len(d["universities"])),
    ]:
        if not os.path.exists(path):
            continue
        docs.append(dict(
            자료명=name, 학년도="-", 발행처="ReBridge 파이프라인",
            경로=os.path.relpath(path, ROOT),
            쪽수="-", 형식=os.path.splitext(path)[1].lstrip(".").upper() or "-",
            텍스트="가능", 수록대학=n, 입수일=mtime(path), 출처URL="",
            용도=use, 비고="원본이 아니라 위 원천에서 만든 산출물"))
    return docs


# ---------------------------------------------------------------------------
# 5. 엑셀 쓰기
# ---------------------------------------------------------------------------
GROUPS = [
    ("기본", ["univId", "대학명", "지역", "설립", "유형", "입학처 URL", "표준코드"]),
    ("2028 시행계획", ["2028 PDF 보유(파일명)", "2028 hwp 원본", "2028 텍스트 추출",
                   "2028 검정고시 판정", "2028 환산표", "2028 원문 게시용 텍스트 준비",
                   "2028 설명회 수록", "2028 반영과목 자료집 수록"]),
    ("2027", ["2027 시행계획", "2027 검정고시 지원가능 전형 수록(권역)", "2027 수록 전형 수",
             "2027 수시 모집요강", "2027 정시 모집요강", "2027 전형분석 자료집 수록",
             "2027 대입정보119 수록", "2027 수시 접수마감 일시"]),
    ("2026", ["2026 전형결과", "2026 수시 모집요강"]),
    ("2025 결과", ["2025 대교협 자료집 수록", "2025 합격선 블록 수", "2025 학과 행 수(DB)"]),
    ("앱 현재 상태", ["앱 전형 행 수", "앱 확인(confirmed)", "앱 템플릿(baseline)만",
                 "앱 합격선 있음", "앱 공식 환산표 반영", "앱 실제 계산 가능(둘 다)"]),
    ("추가 확인 항목", ["정시 수능 반영비율 보유", "모집인원 확보",
                  "검정고시 출신 입학생 수", "대학 자체 입시결과 페이지 URL"]),
    ("", ["비고"]),
]
COLS = [c for _, cs in GROUPS for c in cs]
NUM_COLS = {"2027 수록 전형 수", "2025 합격선 블록 수", "앱 전형 행 수",
            "앱 확인(confirmed)", "앱 템플릿(baseline)만"}
WIDTHS = {"univId": 12, "대학명": 22, "지역": 7, "설립": 8, "유형": 10,
          "입학처 URL": 30, "표준코드": 10, "2028 PDF 보유(파일명)": 42,
          "2028 hwp 원본": 28, "2028 텍스트 추출": 16, "2028 검정고시 판정": 16,
          "2028 환산표": 16, "2028 원문 게시용 텍스트 준비": 15,
          "2028 설명회 수록": 11, "2028 반영과목 자료집 수록": 18,
          "2027 시행계획": 11, "2027 검정고시 지원가능 전형 수록(권역)": 15,
          "2027 수록 전형 수": 10, "2027 수시 모집요강": 12, "2027 정시 모집요강": 12,
          "2027 전형분석 자료집 수록": 14, "2027 대입정보119 수록": 13,
          "2027 수시 접수마감 일시": 17, "2026 전형결과": 11, "2026 수시 모집요강": 13,
          "2025 대교협 자료집 수록": 14, "2025 합격선 블록 수": 12,
          "2025 학과 행 수(DB)": 13, "앱 전형 행 수": 10, "앱 확인(confirmed)": 12,
          "앱 템플릿(baseline)만": 14, "앱 합격선 있음": 11, "앱 공식 환산표 반영": 13,
          "앱 실제 계산 가능(둘 다)": 15, "정시 수능 반영비율 보유": 20,
          "모집인원 확보": 20, "검정고시 출신 입학생 수": 16,
          "대학 자체 입시결과 페이지 URL": 20, "비고": 40}

GROUP_FILL = ["4F6228", "1F4E79", "7030A0", "833C00", "1F6B4E", "5A5A5A", "8B1A1A", "3B3B3B"]
RED = PatternFill("solid", fgColor="FFC7CE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols, group_row=True):
    pass


def write_sheet1(wb, rows):
    ws = wb.create_sheet("대학별 보유현황")
    # 1행: 그룹 헤더
    col = 1
    for gi, (gname, cs) in enumerate(GROUPS):
        if gname:
            ws.cell(row=1, column=col, value=gname)
            if len(cs) > 1:
                ws.merge_cells(start_row=1, start_column=col,
                               end_row=1, end_column=col + len(cs) - 1)
            c = ws.cell(row=1, column=col)
            c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", fgColor=GROUP_FILL[gi % len(GROUP_FILL)])
            c.alignment = Alignment(horizontal="center", vertical="center")
        col += len(cs)
    # 2행: 열 헤더
    for i, name in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(name, 14)
    ws.row_dimensions[2].height = 46
    # 데이터
    for ri, r in enumerate(rows, start=3):
        for ci, name in enumerate(COLS, start=1):
            v = r.get(name, NONE)
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if name in NUM_COLS else "left")
    last = 2 + len(rows)
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{last}"
    rng = f"A3:{get_column_letter(len(COLS))}{last}"
    # 조건부 서식: 없는 것은 빨강, 있는 것은 초록 (빨강을 먼저 걸어 우선 적용)
    for word in ("없음", "판정불가", "미처리", "미수집", "미확인", "검증실패", "마스터에 없음"):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("{word}",A3)))'], fill=RED, stopIfTrue=True))
    for word in ("원문만", "조건부", "확인필요", "미지정", "자료집 제외"):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("{word}",A3)))'], fill=AMBER, stopIfTrue=True))
    for word in ("있음", "가능", "검증통과"):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("{word}",A3)))'], fill=GREEN, stopIfTrue=True))
    # 숫자 0 = 없음
    for name in NUM_COLS:
        L = get_column_letter(COLS.index(name) + 1)
        ws.conditional_formatting.add(
            f"{L}3:{L}{last}", CellIsRule(operator="equal", formula=["0"], fill=RED))
    return ws, last


DOC_COLS = ["자료명", "학년도", "발행처", "파일 경로", "쪽수", "형식", "텍스트 추출 가능",
            "수록 대학 수", "입수일(파일 수정시각)", "출처 URL", "용도", "비고"]
DOC_KEYS = ["자료명", "학년도", "발행처", "경로", "쪽수", "형식", "텍스트",
            "수록대학", "입수일", "출처URL", "용도", "비고"]
DOC_WIDTH = [48, 8, 18, 52, 14, 12, 30, 11, 14, 12, 34, 46]


def write_sheet2(wb, docs):
    ws = wb.create_sheet("자료 목록")
    ws.cell(row=1, column=1,
            value="pdf_sources 아래 원본 자료 + 파이프라인 산출물. "
                  "2028 시행계획 PDF 216개는 한 줄로 묶었다. 출처 URL은 기록이 없어 비워 둔다.")
    ws.cell(row=1, column=1).font = Font(name=FONT, italic=True, size=9, color="808080")
    for i, name in enumerate(DOC_COLS, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = DOC_WIDTH[i - 1]
    for ri, doc in enumerate(docs, start=3):
        for ci, key in enumerate(DOC_KEYS, start=1):
            c = ws.cell(row=ri, column=ci, value=doc.get(key, ""))
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(key in ("자료명", "용도", "비고")))
    last = 2 + len(docs)
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(DOC_COLS))}{last}"
    ws.conditional_formatting.add(
        f"G3:G{last}",
        FormulaRule(formula=['NOT(ISERROR(SEARCH("불가",G3)))'], fill=RED))
    return ws, last


S1 = "'대학별 보유현황'"


def write_sheet3(wb, rows, last1, d):
    ws = wb.create_sheet("요약")
    n4 = sum(1 for r in rows if r.get("유형") == "대학교")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 36
    for L, w in zip("CDEFGH", (10, 13, 13, 12, 12, 62)):
        ws.column_dimensions[L].width = w

    def title(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name=FONT, bold=True, size=13, color="1F4E79")

    def head(row, labels, start=1):
        for i, l in enumerate(labels):
            c = ws.cell(row=row, column=start + i, value=l)
            c.font = Font(name=FONT, bold=True, size=10)
            c.fill = PatternFill("solid", fgColor="D9D9D9")
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.cell(row=1, column=1, value="ReBridge 자료 보유 현황 — 요약")
    ws.cell(row=1, column=1).font = Font(name=FONT, bold=True, size=16)
    ws.cell(row=2, column=1,
            value=f"생성 시각 {datetime.now().strftime('%Y-%m-%d %H:%M')} · "
                  f"대학 {len(rows)}행(마스터 {len(d['universities'])} + 추가 {len(rows)-len(d['universities'])}) · "
                  f"4년제 {n4}개 · 아래 수치는 모두 「대학별 보유현황」을 세는 수식이다")
    ws.cell(row=2, column=1).font = Font(name=FONT, size=9, color="808080")

    title(4, "1. 자료 종류별 커버리지")
    head(5, ["", "자료", "학년도", "대상 대학 수", "보유 대학 수", "미보유", "커버리지", "판정 기준"])

    def colref(name):
        L = get_column_letter(COLS.index(name) + 1)
        return f"{S1}!${L}$3:${L}${last1}"

    KIND_ALL = f"{S1}!$E$3:$E${last1}"
    items = [
        ("2028 시행계획 PDF 원문", "2028", "대학교", "2028 PDF 보유(파일명)", "<>없음",
         "파일명이 적혀 있으면 보유"),
        ("2028 시행계획 hwp 원본", "2028", "대학교", "2028 hwp 원본", "<>없음",
         "PDF가 없거나 깨진 대학의 대체 원본"),
        ("2028 시행계획 텍스트 추출", "2028", "대학교", "2028 텍스트 추출", "있음*",
         "v2/out/text/2028 에 jsonl 이 있으면 보유"),
        ("2028 검정고시 지원가능 판정", "2028", "대학교", "2028 검정고시 판정",
         ["<>판정불가*", "<>미처리"],
         "PDF 근거로 가능/불가/조건부/확인필요 중 하나가 나온 대학"),
        ("2028 검정고시 환산표(구조화 시도)", "2028", "대학교", "2028 환산표", "구조화*",
         "표로 만들어 본 것 전부(검증 실패·앱 미반영 포함)"),
        ("2028 환산표 중 검증통과", "2028", "대학교", "2028 환산표", "구조화·검증통과",
         "등급-점수가 단조라 실제 계산에 쓸 수 있는 상태"),
        ("2028 반영과목 자료집 수록", "2028", "대학교", "2028 반영과목 자료집 수록", "있음",
         "자료집 서두 '발표 대학' 명단"),
        ("2028 대입정보 설명회 수록", "2028", "대학교", "2028 설명회 수록", "있음",
         "설명회 자료 제3부 목차의 대학"),
        ("2027 검정고시 지원가능 전형 목록", "2027", "대학교", "2027 검정고시 지원가능 전형 수록(권역)", "<>없음",
         "5권역 PDF 표에 전형이 1건 이상"),
        ("2027 전형분석 자료집 수록", "2027", "대학교", "2027 전형분석 자료집 수록", "있음",
         "자료집 전문에 대학명 등장"),
        ("2027 대입정보119 수록", "2027", "대학교", "2027 대입정보119 수록", "있음",
         "자료집 전문에 대학명 등장"),
        ("2027 수시 접수마감 일시", "2027", "대학교", "2027 수시 접수마감 일시", "<>없음",
         "접수마감 PDF에서 일시 확보"),
        ("2027 시행계획 원문", "2027", "대학교", "2027 시행계획", "<>없음", "미보유(0)"),
        ("2027 모집요강(수시)", "2027", "대학교", "2027 수시 모집요강", "<>없음", "미보유(0)"),
        ("2026 전형결과", "2026", "대학교", "2026 전형결과", "<>없음", "미보유(0)"),
        ("2026 수시 모집요강", "2026", "대학교", "2026 수시 모집요강", "있음", "성균관대·전북대 2개교뿐"),
        ("2025 전형결과 자료집 수록", "2025", "전체", "2025 대교협 자료집 수록", "있음",
         "지역별 PDF 전문에 대학명 등장"),
        ("2025 합격선 앱 반영", "2025", "전체", "앱 합격선 있음", "있음",
         "cutlines_2025.json 에 항목 존재"),
        ("앱: 확인된 전형 보유", "-", "전체", "앱 확인(confirmed)", ">0",
         "admissions.json status=confirmed*"),
        ("앱: 환산표 반영", "2028", "전체", "앱 공식 환산표 반영", "있음",
         "comparative_2028.json 에 conversion 존재"),
        ("앱: 실제 점수 계산 가능", "-", "전체", "앱 실제 계산 가능(둘 다)", "있음",
         "환산표 + 합격선을 모두 갖춘 대학"),
        ("정시 수능 반영비율(원문이라도)", "2028", "대학교", "정시 수능 반영비율 보유", "원문만",
         "시행계획 텍스트에 '영역별 반영비율' 문구 존재. 구조화된 데이터는 0개"),
        ("모집인원(전형별 인원)", "-", "전체", "모집인원 확보", "<>없음*",
         "recruitCount 가 채워진 전형이 1건 이상"),
        ("검정고시 출신 입학생 수", "-", "전체", "검정고시 출신 입학생 수", "<>미수집", "아직 한 건도 없음"),
        ("대학 자체 입시결과 페이지 URL", "-", "전체", "대학 자체 입시결과 페이지 URL", "<>미확인",
         "아직 한 건도 없음"),
    ]
    r = 6
    for label, yr, scope, col, crit, how in items:
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=yr)
        if scope == "대학교":
            ws.cell(row=r, column=4, value=f'=COUNTIF({KIND_ALL},"대학교")')
        else:
            ws.cell(row=r, column=4, value=f'=COUNTA({colref("대학명")})')
        ref = colref(col)
        crits = crit if isinstance(crit, list) else [crit]
        pairs = "".join(f',{ref},"{c}"' for c in crits)
        if scope == "대학교":
            ws.cell(row=r, column=5, value=f'=COUNTIFS({KIND_ALL},"대학교"{pairs})')
        else:
            ws.cell(row=r, column=5, value=f"=COUNTIFS({pairs[1:]})")
        ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        ws.cell(row=r, column=7, value=f"=IF(D{r}=0,0,E{r}/D{r})")
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=8, value=how)
        for ci in range(2, 9):
            c = ws.cell(row=r, column=ci)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 8),
                                    horizontal="center" if ci in (3, 4, 5, 6, 7) else "left")
        r += 1
    cov_last = r - 1
    ws.conditional_formatting.add(
        f"G6:G{cov_last}",
        CellIsRule(operator="lessThan", formula=["0.2"], fill=RED))
    ws.conditional_formatting.add(
        f"G6:G{cov_last}",
        CellIsRule(operator="between", formula=["0.2", "0.8"], fill=AMBER))
    ws.conditional_formatting.add(
        f"G6:G{cov_last}",
        CellIsRule(operator="greaterThan", formula=["0.8"], fill=GREEN))

    r += 2
    title(r, "2. 다음에 구해올 것 (우선순위)")
    r += 1
    head(r, ["순위", "구해올 자료", "학년도", "결손 대학 수", "", "", "", "왜 필요한가 · 어디서"])
    hdr = r
    r += 1
    # crit 은 '아직 없는 상태'를 세는 조건이다 (커버리지 표와 방향이 반대)
    todo = [
        ("2026학년도 대입 전형결과(대교협 지역별)", "2026", "2026 전형결과", "없음", "대학교",
         "지금 합격선은 2025 한 해뿐이다. 2026이 들어와야 추세를 보고 '작년보다 낮아졌다'를 말할 수 있다. "
         "대교협 어디가/대입정보포털에서 지역별 PDF로 배포된다"),
        ("2027학년도 대학입학전형 시행계획", "2027", "2027 시행계획", "없음", "대학교",
         "2027 입시를 치르는 학생에게 지금 줄 수 있는 게 5권역 지원자격표뿐이다. "
         "전형방법·수능최저·환산표가 없다. 각 대학 입학처 또는 대교협 취합본"),
        ("2027 수시 모집요강(최소 상위 30개교)", "2027", "2027 수시 모집요강", "없음", "대학교",
         "시행계획보다 확정적이다. 실제 지원 판단은 모집요강 기준이어야 한다. 각 대학 입학처"),
        ("2027 정시 모집요강 / 정시 수능 반영비율", "2027", "2027 정시 모집요강", "없음", "대학교",
         "검정고시생 상당수가 정시를 본다. 지금 정시 반영비율은 구조화된 게 0개다"),
        ("검정고시 출신 입학생 수(대학알리미 4-자 공시)", "-", "검정고시 출신 입학생 수", "미수집", "전체",
         "'이 대학이 검정고시생을 실제로 뽑는가'를 숫자로 보여줄 유일한 자료. "
         "대학알리미(academyinfo.go.kr) 공시 4-자 항목을 대학별로 내려받아 붙인다"),
        ("대학 자체 입시결과 페이지 URL", "-", "대학 자체 입시결과 페이지 URL", "미확인", "전체",
         "대교협 자료집보다 상세한 학과별 결과가 대학 입학처 홈페이지에 있다. "
         "우선 4년제부터 입학처 URL을 타고 '입시결과' 페이지를 수집한다"),
        ("2028 시행계획 텍스트 추출 실패분 복구", "2028", "2028 텍스트 추출", "없음*", "대학교",
         "PDF는 있는데 텍스트가 안 뽑힌 대학. hwp 원본 또는 OCR로 복구"),
        ("2028 환산표 구조화(원문만 상태 해소)", "2028", "2028 환산표", "원문만", "대학교",
         "원문은 있는데 표로 못 만든 상태. 이게 풀려야 점수 계산이 된다"),
        ("대학 표준코드(공공데이터 학교코드)", "-", "표준코드", "없음", "전체",
         "지금 univId가 uA…/슬러그 혼재라 외부 데이터와 붙일 수 없다. 교육부 학교기본정보 표준코드"),
    ]
    for i, (label, yr, col, crit, scope, why) in enumerate(todo, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=yr)
        ref = colref(col)
        if scope == "대학교":
            ws.cell(row=r, column=4,
                    value=f'=COUNTIFS({KIND_ALL},"대학교",{ref},"{crit}")')
        else:
            ws.cell(row=r, column=4, value=f'=COUNTIF({ref},"{crit}")')
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        ws.cell(row=r, column=8, value=why)
        for ci in list(range(1, 5)) + [8]:
            c = ws.cell(row=r, column=ci)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 8),
                                    horizontal="center" if ci in (1, 3, 4) else "left")
        ws.row_dimensions[r].height = 34
        r += 1
    ws.cell(row=hdr, column=4, value="결손 대학 수")

    r += 2
    title(r, "3. 추출이 실패했거나 확실하지 않은 것")
    r += 1
    for w in (warnings or ["없음"]):
        c = ws.cell(row=r, column=2, value="· " + w)
        c.font = Font(name=FONT, size=10, color="9C0006")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


EXPLAIN = [
    # (시트, 열, 뜻, 판정 기준, 원천)
    ("대학별 보유현황", "univId", "앱 내부 대학 식별자", "universities.json 의 univId. 빈칸이면 마스터에 없는 대학", "src/data/universities.json"),
    ("대학별 보유현황", "대학명 / 지역 / 설립 / 유형", "대학 기본 정보", "마스터 그대로. 유형은 대학교(4년제)/전문대학", "src/data/universities.json"),
    ("대학별 보유현황", "입학처 URL", "대학 입학처 주소", "마스터 그대로", "src/data/universities.json"),
    ("대학별 보유현황", "표준코드", "공공데이터 학교 표준코드", "DB university.std_code. 현재 351개 전부 비어 있어 '없음'", "rebridge.db"),
    ("대학별 보유현황", "2028 PDF 보유(파일명)", "2028 시행계획 원본 PDF", "pdf_sources/2028 파일명. 캠퍼스가 여러 개면 ' / '로 이어 붙였다", "pdf_sources/2028"),
    ("대학별 보유현황", "2028 hwp 원본", "PDF 대신·함께 있는 hwp/hwpx", "pdf_sources/2028_hwp 파일명", "pdf_sources/2028_hwp"),
    ("대학별 보유현황", "2028 텍스트 추출", "PDF에서 글자를 뽑아 뒀는가", "v2/out/text/2028/{univId}*.jsonl 존재 여부와 쪽수", "v2/out/text/2028"),
    ("대학별 보유현황", "2028 검정고시 판정", "시행계획 원문 기준 검정고시 지원 가능 여부",
     "plans_2028/ged.jsonl 의 eligible. 가능/불가/조건부/확인필요 = 원문에서 근거를 찾은 것, "
     "'판정불가(근거 문구 못 찾음)' = PDF는 읽었으나 지원자격 문구를 못 찾은 것, "
     "'판정불가(원문 없음)' = 그 대학 2028 시행계획 자체가 없는 것, "
     "'미처리' = PDF는 있는데 판정을 안 돌린 것. 캠퍼스별로 결과가 다르면 "
     "가능>조건부>확인필요>불가>판정불가 순으로 하나를 적고 '캠퍼스별 상이'를 붙였다",
     "plans_2028/ged.jsonl"),
    ("대학별 보유현황", "2028 환산표", "검정고시 성적을 학생부 점수로 바꾸는 표의 상태",
     "구조화·검증통과 = 표로 만들었고 등급-점수가 단조(monotonic) / 구조화·검증실패 = 표는 있으나 값이 뒤집힘 "
     "/ 구조화·앱 미반영 = 파이프라인에만 있고 앱 JSON에 없음 / 원문만 = 글만 있고 표 없음 / 없음",
     "comparative_2028.json + plans_2028/conversion.jsonl"),
    ("대학별 보유현황", "2028 원문 게시용 텍스트 준비", "앱에서 '근거 원문'을 보여줄 문장이 확보됐는가",
     "ged.jsonl 의 quote 또는 comparative_2028.json 의 comparativeGrade 가 비어 있지 않으면 있음",
     "plans_2028/ged.jsonl, comparative_2028.json"),
    ("대학별 보유현황", "2028 설명회 수록", "2028 대입 정보 설명회 자료 제3부 수록 여부",
     "설명회 PDF가 이미지라 본문 검색이 안 된다. 목차 제3부에 이름이 있는 7개 대학만 '있음'",
     "kcue_2028/2028 대입 정보 설명회"),
    ("대학별 보유현황", "2028 반영과목 자료집 수록", "권장과목 자료집에 실렸는가",
     "자료집 서두 명단 그대로. 있음(발표 대학) / 권장과목 미지정 / 자료집 제외(학종 미실시) / 없음",
     "kcue_2028/모집단위별 반영과목 자료집"),
    ("대학별 보유현황", "2027 시행계획", "2027 시행계획 원문", "한 건도 없다. 전부 '없음'", "-"),
    ("대학별 보유현황", "2027 검정고시 지원가능 전형 수록(권역)", "5권역 PDF 중 어디에 실렸는가",
     "pdfplumber 로 표를 뽑아 대학명으로 매칭한 권역명", "pdf_sources/ged_eligible_2027"),
    ("대학별 보유현황", "2027 수록 전형 수", "그 대학의 검정고시 지원가능 전형 개수",
     "5권역 PDF 표에서 (지역·대학·전형명)이 채워진 행 수. 0 = 없음", "pdf_sources/ged_eligible_2027"),
    ("대학별 보유현황", "2027 수시/정시 모집요강", "2027 모집요강 원문", "한 건도 없다. 전부 '없음'", "-"),
    ("대학별 보유현황", "2027 전형분석 자료집 수록", "지역별 전형분석 자료집(728쪽)에 나오는가",
     "PDF 전문 텍스트에 대학명(전체명 또는 '고려대' 같은 약칭)이 등장하면 있음. "
     "쪽 단위 검증은 하지 않았으므로 대략적인 판정이다", "kcue_2027/지역별 전형분석 자료집"),
    ("대학별 보유현황", "2027 대입정보119 수록", "대입정보119 자료집(654쪽)에 나오는가", "위와 같은 방식", "kcue_2027/대입정보 119 자료집"),
    ("대학별 보유현황", "2027 수시 접수마감 일시", "원서접수 마감 날짜·시각",
     "접수마감 PDF(3쪽)에서 지역-일자-시각-대학 목록을 파싱. 줄바꿈으로 잘린 대학명은 이어 붙였다",
     "kcue_2027/수시모집 전형 일정(접수 마감)"),
    ("대학별 보유현황", "2026 전형결과", "2026 대입 전형결과", "한 건도 없다. 전부 '없음'", "-"),
    ("대학별 보유현황", "2026 수시 모집요강", "2026 수시 모집요강 PDF", "guides_2026 폴더. 성균관대·전북대 2개교뿐", "pdf_sources/guides_2026"),
    ("대학별 보유현황", "2025 대교협 자료집 수록", "2025 전형결과 지역별 PDF에 이름이 나오는가",
     "34개 PDF 전문 텍스트에 대학명이 등장하면 있음. 일부 PDF는 이미지라 글자가 거의 안 뽑혔다(「자료 목록」 참고)",
     "pdf_sources/2025"),
    ("대학별 보유현황", "2025 합격선 블록 수", "앱이 쓰는 2025 합격선 묶음 수",
     "cutlines_2025.json 의 전형유형 키 개수(논술/학생부교과/학생부종합 등). 0 = 없음", "src/data/cutlines_2025.json"),
    ("대학별 보유현황", "2025 학과 행 수(DB)", "DB에 들어간 2025 합격선 원자료 행 수",
     "rebridge.db cutline 테이블 year=2025 행 수", "rebridge.db"),
    ("대학별 보유현황", "앱 전형 행 수", "앱이 이 대학에 대해 가진 전형 개수", "admissions.json 행 수", "src/data/admissions.json"),
    ("대학별 보유현황", "앱 확인(confirmed)", "원문 근거로 확인된 전형 수",
     "status 가 confirmed / confirmed_detail 인 행", "src/data/admissions.json"),
    ("대학별 보유현황", "앱 템플릿(baseline)만", "근거 없이 기본값으로 채워 둔 전형 수",
     "status=baseline. 이 숫자가 크면 그 대학 화면은 사실상 '일반론'만 보여 준다", "src/data/admissions.json"),
    ("대학별 보유현황", "앱 합격선 있음", "합격선을 화면에 띄울 수 있는가", "cutlines_2025.json 에 해당 univId 존재", "src/data/cutlines_2025.json"),
    ("대학별 보유현황", "앱 공식 환산표 반영", "검정고시 점수 환산표가 앱에 들어갔는가",
     "comparative_2028.json 에 conversion 키 존재", "src/data/comparative_2028.json"),
    ("대학별 보유현황", "앱 실제 계산 가능(둘 다)", "'내 점수로 되나?'를 계산해 줄 수 있는가",
     "환산표와 합격선을 모두 갖춘 대학만 '있음'", "comparative_2028.json + cutlines_2025.json"),
    ("대학별 보유현황", "정시 수능 반영비율 보유", "정시 수능 영역별 반영비율 확보 상태",
     "2028 시행계획 추출 텍스트에 '영역별 반영비율' 등의 문구가 있으면 '원문만'. "
     "표로 구조화된 데이터는 아직 한 건도 없어 '있음'은 나오지 않는다",
     "v2/out/text/2028"),
    ("대학별 보유현황", "모집인원 확보", "전형별 모집인원을 알고 있는가",
     "'recruitCount 있는 전형 수 / 전체 전형 수 (DB admission_program 행 수)'. "
     "DB admission_program 은 현재 0행이다", "src/data/admissions.json + rebridge.db"),
    ("대학별 보유현황", "검정고시 출신 입학생 수", "그 대학이 실제로 검정고시생을 몇 명 뽑았는가",
     "대학알리미 4-자 공시 항목. 아직 수집 전이라 전부 '미수집'", "(미수집)"),
    ("대학별 보유현황", "대학 자체 입시결과 페이지 URL", "대학 입학처의 입시결과 페이지 주소",
     "아직 수집 전이라 전부 '미확인'", "(미확인)"),
    ("대학별 보유현황", "비고", "그 행에서 사람이 알아야 할 특이사항",
     "캠퍼스 PDF 합침 / 텍스트 추출 실패 / 마스터에 없음 등", "-"),
    ("자료 목록", "쪽수", "PDF 쪽수", "pdfinfo 결과. hwp 는 '-'", "-"),
    ("자료 목록", "텍스트 추출 가능", "글자를 뽑을 수 있는가",
     "pdftotext 결과 길이로 판단. 5,000자 미만이면 이미지 기반으로 보고 '부분/불가'", "-"),
    ("자료 목록", "수록 대학 수", "그 자료가 다루는 대학 수",
     "자료마다 기준이 다르다. 5권역 PDF는 표에 나온 대학 수, 대형 자료집은 전문에 이름이 나온 대학 수", "-"),
    ("자료 목록", "입수일(파일 수정시각)", "파일이 디스크에 기록된 날",
     "실제 발간일이 아니라 파일 mtime 이다. 원본 발간일은 따로 확인이 필요하다", "-"),
    ("자료 목록", "출처 URL", "어디서 받았는가", "기록이 남아 있지 않아 전부 비어 있다. 다음 수집 때부터 채울 것", "-"),
    ("요약", "대상 대학 수", "그 자료가 원래 커버해야 할 대학 수",
     "시행계획·자료집류는 4년제(대학교)만 대상이라 4년제 수를 쓴다. 나머지는 전체 행 수", "-"),
    ("요약", "커버리지", "보유 / 대상",
     "20% 미만 빨강, 20~80% 노랑, 80% 초과 초록", "-"),
    ("요약", "결손 대학 수", "아직 못 구한 대학 수",
     "「대학별 보유현황」을 COUNTIFS 로 센 값이다. 원본 시트가 바뀌면 자동으로 따라 바뀐다", "-"),
]


def write_sheet4(wb):
    ws = wb.create_sheet("열 설명")
    ws.cell(row=1, column=1, value="열 설명 — 각 칸의 뜻과 어떻게 판정했는지")
    ws.cell(row=1, column=1).font = Font(name=FONT, bold=True, size=14)
    ws.cell(row=2, column=1,
            value="'없음'은 정말 없다는 뜻이다. 모르는 것은 '미수집'/'미확인'/'판정불가'로 따로 적었다.")
    ws.cell(row=2, column=1).font = Font(name=FONT, size=9, color="808080")
    heads = ["시트", "열", "무슨 뜻인가", "어떻게 판정했나", "원천"]
    for i, h in enumerate(heads, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name=FONT, bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    for L, w in zip("ABCDE", (16, 30, 40, 74, 34)):
        ws.column_dimensions[L].width = w
    for ri, row in enumerate(EXPLAIN, start=4):
        for ci, v in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{3 + len(EXPLAIN)}"
    return ws


# ---------------------------------------------------------------------------
def main():
    d = collect()
    rows = build_rows(d)
    docs = build_docs(d, rows)

    wb = Workbook()
    wb.remove(wb.active)
    ws1, last1 = write_sheet1(wb, rows)
    ws2, last2 = write_sheet2(wb, docs)
    write_sheet3(wb, rows, last1, d)
    write_sheet4(wb)
    wb.move_sheet("요약", offset=-2)          # 요약을 앞으로
    wb.save(OUT_XLSX)
    log(f"\n저장: {OUT_XLSX}")
    log(f"  대학별 보유현황: {len(rows)}행 x {len(COLS)}열")
    log(f"  자료 목록: {len(docs)}행 x {len(DOC_COLS)}열")
    log(f"  열 설명: {len(EXPLAIN)}행 x 5열")

    stats = {
        "rows": len(rows), "cols": len(COLS), "docs": len(docs),
        "zone_stats": d["zone_stats"],
        "ged27_univ_names": len(d["ged27_names"]),
        "ged27_unmatched": dict(d["ged27_unmatched"]),
        "extra_rows": d["extra_rows"],
        "no_text_2028": d["no_text_2028"],
        "thin_2025": d["thin_2025"],
        "warnings": warnings,
        "deadline_matched": len(d["deadline"]),
        "in_analysis": len(d["in_analysis"]), "in_119": len(d["in_119"]),
        "in_2025": len(d["in_2025"]), "in_briefing": sorted(d["in_briefing"]),
        "subjects": {k: len(v) for k, v in d["subjects"].items() if k != "raw"},
    }
    json.dump(stats, open(os.path.join(HERE, "build_stats.json"), "w"),
              ensure_ascii=False, indent=1)
    log("통계: build_stats.json")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--refresh", action="store_true", help="PDF 추출 캐시를 다시 만든다")
    args = ap.parse_args()
    main()
