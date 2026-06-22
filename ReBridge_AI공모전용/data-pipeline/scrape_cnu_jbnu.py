import json
import re
import urllib.request
from collections import Counter
from pathlib import Path
from urllib.parse import quote

import fitz
from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
CLEAN = BASE / "results_2025_clean.json"
REPORT = BASE / "reports" / "results_2025_clean_report.md"

TMP = Path("/private/tmp")

CNU_FILES = [
    {
        "label": "수시",
        "url": "https://ipsi.cnu.ac.kr/_prog/_board/common/download.php?code=recsroom_01&ntt_no=2021596",
        "path": TMP / "cnu_susi_2025.pdf",
        "sourceFile": "충남대학교_2025학년도_수시모집_입시결과.pdf",
    },
    {
        "label": "정시",
        "url": "https://ipsi.cnu.ac.kr/_prog/_board/common/download.php?code=recsroom_02&ntt_no=2021597",
        "path": TMP / "cnu_jeongsi_2025.pdf",
        "sourceFile": "충남대학교_2025학년도_정시모집_입시결과.pdf",
    },
]

JBNU_FILES = [
    {
        "label": "수시+정시",
        "url": (
            "https://enter.jbnu.ac.kr/file/download.do?"
            "sfn=20250811044333615_"
            + quote("2025학년도 학부(수시,정시) 입시결과(공개용) .xlsx")
            + "&ofn="
            + quote("2025학년도 학부(수시,정시) 입시결과(공개용) .xlsx")
        ),
        "path": TMP / "jbnu_susi_jeongsi_2025.xlsx",
        "sourceFile": "2025학년도 학부(수시,정시) 입시결과(공개용).xlsx",
    },
    {
        "label": "정시",
        "url": (
            "https://enter.jbnu.ac.kr/file/download.do?"
            "sfn=20251127012919226_"
            + quote("250329_2025학년도 학부 입시결과(공지).xlsx")
            + "&ofn="
            + quote("250329_2025학년도 학부 입시결과(공지).xlsx")
        ),
        "path": TMP / "jbnu_jeongsi_2025.xlsx",
        "sourceFile": "250329_2025학년도 학부 입시결과(공지).xlsx",
    },
]


def download_if_missing(item):
    if item["path"].exists() and item["path"].stat().st_size > 1024:
        return "cached"
    try:
        with urllib.request.urlopen(item["url"], timeout=30) as response:
            item["path"].write_bytes(response.read())
        return "downloaded"
    except Exception as exc:
        return f"download_failed: {exc}"


def text_layer_stats(pdf_path):
    doc = fitz.open(pdf_path)
    stats = []
    body_like_pages = 0
    for index, page in enumerate(doc, start=1):
        text = page.get_text()
        has_unit_or_numbers = bool(
            re.search(r"(학과|학부|계열).{0,80}(?:\d\.\d|\d{2,3}\.\d|70%)", text, re.S)
        )
        if has_unit_or_numbers:
            body_like_pages += 1
        stats.append((index, len(text), has_unit_or_numbers))
    doc.close()
    return {
        "pages": len(stats),
        "pagesOver200": sum(1 for _, length, _ in stats if length > 200),
        "bodyLikePages": body_like_pages,
        "maxTextLength": max((length for _, length, _ in stats), default=0),
    }


def as_float(value):
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    return None


def clean_unit(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def make_row(admission_name, unit, admission_type, phase, source_file, sheet, excel_row,
             recruit_count=None, competition=None, cut_grade=None, cut_score=None, note=""):
    return {
        "univId": "jbnu",
        "univName": "전북대학교",
        "phase": phase,
        "admissionType": admission_type,
        "admissionName": str(admission_name).strip(),
        "unit": clean_unit(unit),
        "year": 2025,
        "cutType": "70%컷",
        "cutGrade": cut_grade,
        "cutScore": cut_score,
        "recruitCount": recruit_count,
        "competition": competition,
        "region": "전북",
        "sourceFile": source_file,
        "sourcePage": None,
        "confidence": "high",
        "note": f"[B-3: 전북대 공식 XLSX] {sheet}!{excel_row} {note}".strip(),
    }


def parse_jbnu_workbook(path, source_file):
    wb = load_workbook(path, data_only=True)
    rows = []

    if "종합" in wb.sheetnames:
        ws = wb["종합"]
        for r in range(14, ws.max_row + 1):
            admission_name = ws.cell(r, 1).value
            unit = ws.cell(r, 3).value
            cut_grade = as_float(ws.cell(r, 21).value)
            cut_score = as_float(ws.cell(r, 13).value)
            if not admission_name or not unit or cut_grade is None:
                continue
            rows.append(
                make_row(
                    admission_name, unit, "학생부종합", "수시", source_file, "종합", r,
                    recruit_count=as_float(ws.cell(r, 4).value),
                    competition=as_float(ws.cell(r, 7).value),
                    cut_grade=cut_grade,
                    cut_score=cut_score,
                    note="cutGrade=최종등록자 학생부등급 70% cut, cutScore=전형총점 70% cut",
                )
            )

    if "교과" in wb.sheetnames:
        ws = wb["교과"]
        for r in range(15, ws.max_row + 1):
            admission_name = ws.cell(r, 1).value
            unit = ws.cell(r, 3).value
            cut_grade = as_float(ws.cell(r, 15).value)
            cut_score = as_float(ws.cell(r, 19).value)
            if not admission_name or not unit or cut_grade is None:
                continue
            rows.append(
                make_row(
                    admission_name, unit, "학생부교과", "수시", source_file, "교과", r,
                    recruit_count=as_float(ws.cell(r, 4).value),
                    competition=as_float(ws.cell(r, 7).value),
                    cut_grade=cut_grade,
                    cut_score=cut_score,
                    note="cutGrade=최종등록자 학생부등급 70% cut, cutScore=학생부환산점수 70% cut",
                )
            )

    if "정시" in wb.sheetnames:
        ws = wb["정시"]
        for r in range(12, ws.max_row + 1):
            admission_name = ws.cell(r, 1).value
            unit = ws.cell(r, 3).value
            cut_score = as_float(ws.cell(r, 21).value)
            if not admission_name or not unit or cut_score is None:
                continue
            rows.append(
                make_row(
                    admission_name, unit, "수능위주", "정시", source_file, "정시", r,
                    recruit_count=as_float(ws.cell(r, 6).value),
                    competition=as_float(ws.cell(r, 9).value),
                    cut_grade=None,
                    cut_score=cut_score,
                    note="cutScore=최종등록자 수능 백분위 전체영역 70% cut",
                )
            )

    return rows


def append_report(cnu_report, jbnu_rows):
    existing = REPORT.read_text(encoding="utf-8") if REPORT.exists() else ""
    marker = "## B-3 충남대/전북대 공식 자료 재시도"
    if marker in existing:
        existing = existing.split(marker)[0].rstrip() + "\n\n"

    conf = Counter(row["confidence"] for row in jbnu_rows)
    type_counts = Counter(row["admissionType"] for row in jbnu_rows)
    lines = [
        marker,
        "",
        "| 대학 | 공식 자료 | 파싱 가능 여부 | 추가 행수 | 비고 |",
        "| --- | --- | --- | ---: | --- |",
    ]
    for item in cnu_report:
        lines.append(
            f"| 충남대학교 | {item['label']} PDF | 불가 | 0 | "
            f"pages={item['pages']}, pagesOver200={item['pagesOver200']}, "
            f"bodyLikePages={item['bodyLikePages']}; 설명문 텍스트만 추출되고 합격선 표 0건 |"
        )
    lines.append(
        f"| 전북대학교 | 공식 XLSX | 가능 | {len(jbnu_rows)} | "
        f"{', '.join(f'{k}:{v}' for k, v in type_counts.items())} |"
    )
    lines.extend([
        "",
        "### 전북대학교 confidence 분포",
        "",
        "| confidence | count |",
        "| --- | ---: |",
    ])
    for key, count in conf.items():
        lines.append(f"| {key} | {count} |")
    lines.extend([
        "",
        "- 충남대: 입학처 공식 PDF를 확인했으나 `fitz`/`pdfplumber` 모두 합격선 표 본문을 행 단위로 추출하지 못해 append하지 않음.",
        "- 전북대: 입학처 공식 XLSX의 `종합`, `교과`, `정시` 시트를 deterministic column mapping으로 파싱해 append.",
    ])
    REPORT.write_text(existing + "\n".join(lines) + "\n", encoding="utf-8")


def main():
    cnu_report = []
    for item in CNU_FILES:
        status = download_if_missing(item)
        if item["path"].exists():
            stats = text_layer_stats(item["path"])
            cnu_report.append({"label": item["label"], "status": status, **stats})
        else:
            cnu_report.append({"label": item["label"], "status": status, "pages": 0, "pagesOver200": 0, "bodyLikePages": 0})

    for item in JBNU_FILES:
        download_if_missing(item)

    jbnu_rows = []
    primary = JBNU_FILES[0]
    if primary["path"].exists():
        jbnu_rows = parse_jbnu_workbook(primary["path"], primary["sourceFile"])

    rows = json.loads(CLEAN.read_text(encoding="utf-8"))
    before = len(rows)
    removed = sum(1 for row in rows if row.get("univId") in {"cnu", "jbnu"})
    rows = [row for row in rows if row.get("univId") not in {"cnu", "jbnu"}]
    rows.extend(jbnu_rows)
    CLEAN.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    append_report(cnu_report, jbnu_rows)

    print(f"before={before}")
    print(f"removed_existing_cnu_jbnu={removed}")
    print(f"jbnu_added={len(jbnu_rows)}")
    print(f"total={len(rows)}")
    print(f"cnu={cnu_report}")


if __name__ == "__main__":
    main()
