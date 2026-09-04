#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""대학알리미 '신입생의 출신 고등학교 유형별 현황'(4년제) XLSX → 검정고시 입학생 수 추출.

입력 : Application_main_codes/src/data/pdf_sources/academyinfo/4자_출신고교유형_{2024,2025,2026}.xlsx
출력 : data-pipeline/v2/out/academyinfo/ged_freshmen.jsonl  (한 줄 = 대학 × 연도)

시트 구조(3개 연도 모두 동일, 실측):
  1~3행  제목/공백
  4~9행  다단 헤더 (병합셀)
  10행~  데이터
  열(0-base):
    0 기준연도 / 1 학교종류 / 2 설립구분 / 3 지역 / 4 상태 / 5 학교 / 6 총입학자수
    7~34  출신학교 유형별 (학생수,비율) 쌍 … 29=검정고시 학생수, 30=검정고시 비율, 33=소계
    35~48 출신학교 지역정보별
  검증: 7·9·11·13·15·17·19·21·23·25·27·29·31 합 = 33(소계) = 6(총입학자수)

주의: 이 XLSX는 colDimension 속성명이 'cumstomWidth'로 잘못 적혀 있어
      openpyxl 일반 모드로는 못 연다. read_only + reset_dimensions()로 읽는다.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import PDF_ROOT, OUT, UnivMatcher, squash  # noqa: E402

SRC_DIR = PDF_ROOT / "academyinfo"
OUT_DIR = OUT / "academyinfo"
YEARS = (2024, 2025, 2026)

HEADER_ROWS = 9          # 10행부터 데이터
C_YEAR, C_KIND, C_ESTAB, C_REGION, C_STATUS, C_NAME, C_TOTAL = 0, 1, 2, 3, 4, 5, 6
C_GED_N, C_GED_PCT, C_SUBTOTAL = 29, 30, 33

# 유형별 (라벨, 학생수 열) — 합계 검증용
TYPE_COLS = [
    ("일반고", 7), ("과학고", 9), ("외고국제고", 11), ("예술체육고", 13),
    ("산업수요맞춤형고", 15), ("특성화고", 17), ("자율고(사립)", 19), ("자율고(공립)", 21),
    ("영재학교", 23), ("외국인학교", 25), ("외국고등학교", 27), ("검정고시", 29),
    ("그외기타", 31),
]


# 알리미는 캠퍼스 행을 '강원대학교_제2캠퍼스', '연세대학교(미래)_분교'처럼 접미사로 구분한다.
# universities.json에는 이 접미사가 없으므로 떼고 매칭한다.
_CAMPUS_SUFFIX = re.compile(r"_(제\d+캠퍼스|분교|본교)\s*$")


def split_campus(name: str):
    """'연세대학교(미래)_분교' → ('연세대학교(미래)', '분교')"""
    m = _CAMPUS_SUFFIX.search(name)
    if not m:
        return name, "본교"
    return name[: m.start()].strip(), m.group(1)


def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return int(f) if f == int(f) else f


def read_sheet(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()          # 시트의 dimension ref가 'A1'로 깨져 있어 필수
    rows = [r for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    matcher = UnivMatcher()

    records = []
    unmatched = []          # (year, name, region, kind)
    stats = {}

    for year in YEARS:
        path = SRC_DIR / f"4자_출신고교유형_{year}.xlsx"
        if not path.exists():
            raise SystemExit(f"입력 파일 없음: {path}")

        rows = read_sheet(path)
        n_row = n_ok = n_bad_sum = 0

        for r in rows[HEADER_ROWS:]:
            if not r or len(r) <= C_SUBTOTAL:
                continue
            name = squash(r[C_NAME]) if r[C_NAME] else ""
            if not name or not r[C_YEAR]:
                continue
            n_row += 1

            region = squash(r[C_REGION])
            total = num(r[C_TOTAL])
            ged_n = num(r[C_GED_N])
            ged_pct = num(r[C_GED_PCT])
            subtotal = num(r[C_SUBTOTAL])

            # 열 위치 검증 — 유형별 합이 소계/총입학자수와 맞는지
            parts = [num(r[c]) or 0 for _, c in TYPE_COLS]
            sum_ok = (sum(parts) == subtotal == total)
            if not sum_ok:
                n_bad_sum += 1

            base, campus = split_campus(name)
            uid = matcher.match(base, region=region)
            if uid:
                n_ok += 1
            else:
                unmatched.append((year, name, region, squash(r[C_KIND])))

            records.append({
                "year": int(float(r[C_YEAR])),
                "univId": uid,
                "univName": matcher.name_of(uid) if uid else None,
                "sourceName": name,
                "campus": campus,
                "schoolKind": squash(r[C_KIND]),
                "establishment": squash(r[C_ESTAB]),
                "region": region,
                "status": squash(r[C_STATUS]),
                "totalFreshmen": total,
                "gedFreshmen": ged_n,
                "gedPercent": ged_pct,
                "typeSubtotal": subtotal,
                "sumCheck": sum_ok,
                "matched": bool(uid),
                "source": f"academyinfo/4자_출신고교유형_{year}.xlsx",
            })

        stats[year] = {"rows": n_row, "matched": n_ok,
                       "unmatched": n_row - n_ok, "sumMismatch": n_bad_sum}

    out_path = OUT_DIR / "ged_freshmen.jsonl"
    with open(out_path, "w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # 파생본: univId × 연도로 캠퍼스 행을 합산한 것 (앱이 바로 쓰기 좋은 형태)
    agg = {}
    for r in records:
        if not r["univId"]:
            continue
        k = (r["univId"], r["year"])
        a = agg.setdefault(k, {
            "univId": r["univId"], "univName": r["univName"], "year": r["year"],
            "region": r["region"], "totalFreshmen": 0, "gedFreshmen": 0,
            "campusRows": [], "source": r["source"],
        })
        a["totalFreshmen"] += r["totalFreshmen"] or 0
        a["gedFreshmen"] += r["gedFreshmen"] or 0
        a["campusRows"].append(r["sourceName"])
    for a in agg.values():
        a["gedPercent"] = round(a["gedFreshmen"] / a["totalFreshmen"] * 100, 2) if a["totalFreshmen"] else None
    agg_path = OUT_DIR / "ged_freshmen_by_univ.jsonl"
    with open(agg_path, "w", encoding="utf-8") as f:
        for k in sorted(agg):
            f.write(json.dumps(agg[k], ensure_ascii=False) + "\n")

    # 리포트
    lines = ["# 대학알리미 검정고시 입학생 추출 리포트", ""]
    lines.append("| 연도 | 데이터 행 | univId 매칭 | 매칭 실패 | 합계검증 실패 |")
    lines.append("|---|---|---|---|---|")
    for y in YEARS:
        s = stats[y]
        lines.append(f"| {y} | {s['rows']} | {s['matched']} | {s['unmatched']} | {s['sumMismatch']} |")
    # 같은 univId가 한 연도에 여러 행으로 잡히는 경우(본교 + 제2캠퍼스)를 드러낸다.
    dup = Counter((r["year"], r["univId"]) for r in records if r["univId"])
    dups = sorted(k for k, v in dup.items() if v > 1)

    lines += ["", f"총 레코드: {len(records)}", ""]
    lines += ["## univId 중복 (본교 + 캠퍼스 행이 같은 univId로 매칭됨)", "",
              "알리미는 캠퍼스를 별도 행으로 공시하지만 universities.json에는 대응 항목이 없다.",
              "합산해서 쓰려면 소비 측에서 univId+year로 sum 해야 한다. 그냥 쓰면 이중계상된다.", ""]
    if not dups:
        lines.append("- 없음")
    else:
        lines.append("| 연도 | univId | 대학 | 행 수 | 원본 표기 |")
        lines.append("|---|---|---|---|---|")
        for y, uid in dups:
            src = [r["sourceName"] for r in records if r["year"] == y and r["univId"] == uid]
            lines.append(f"| {y} | {uid} | {matcher.name_of(uid)} | {len(src)} | {' / '.join(src)} |")
    # 반대 방향: universities.json의 4년제 중 알리미 공시에 아예 없는 학교
    four = {u["univId"]: u["name"] for u in matcher.univs if u.get("kind") == "대학교"}
    have = {r["univId"] for r in records if r["univId"]}
    missing = sorted(((i, n) for i, n in four.items() if i not in have), key=lambda x: x[1])
    lines += ["", "## universities.json 4년제 중 대학알리미 공시에 없는 학교", "",
              "대학알리미는 고등교육법상 정규 4년제를 전수 공시한다. 여기 빠졌다는 것은",
              "사내대학·각종학교이거나 마스터 데이터 오류일 가능성을 뜻한다.", "",
              f"4년제 마스터 {len(four)}개 중 {len(four) - len(missing)}개 매칭, {len(missing)}개 미포함", ""]
    for i, n in missing:
        lines.append(f"- {n} ({i})")
    lines += ["", "## 매칭 실패 대학 (연도별 전체 나열)", ""]
    for y in YEARS:
        us = [u for u in unmatched if u[0] == y]
        lines.append(f"### {y} — {len(us)}개")
        if not us:
            lines.append("- 없음")
        for _, nm, rg, kd in us:
            lines.append(f"- {nm} ({rg} / {kd})")
        lines.append("")

    (OUT_DIR / "ged_freshmen_report.md").write_text("\n".join(lines), encoding="utf-8")

    print(f"wrote {out_path} ({len(records)} records)")
    print(f"wrote {agg_path} ({len(agg)} univ×year)")
    for y in YEARS:
        print(" ", y, stats[y])
    uniq = sorted({(nm, rg, kd) for _, nm, rg, kd in unmatched})
    print(f"unmatched unique names: {len(uniq)}")
    for nm, rg, kd in uniq:
        print("   -", nm, "|", rg, "|", kd)


if __name__ == "__main__":
    main()
